#!/usr/bin/env python3
"""
從 Excel 解析遊戲資料並生成 JSON

功能：
1. 解析農作物資料（處理變異種繼承）
2. 解析風車加工品資料
3. 解析料理資料
4. 自動處理季節、材料等欄位
"""

import openpyxl
import json
import re
from pathlib import Path

# 變異種前綴
VARIANT_PREFIXES = ['巨大', '扭曲', '結實', '黃金', '波浪', '葫蘆形', '愛心', '珠寶',
                    '紫', '心形', '指甲', '三腳', '雙腳', '跳舞', '萬歲', '群花',
                    '彩色', '圓球', '雞型', '腫包']

# 特殊命名的變異種對應表
VARIANT_MAPPING = {
    '星形馬鈴薯': '馬鈴薯',
    '愛心西瓜': '西瓜',
    '珠寶哈密瓜': '哈密瓜',
    '三腳胡蘿蔔': '胡蘿蔔',
    '愛心菠菜': '菠菜',
    '指甲辣椒': '辣椒',
    '黃金青江菜': '青江菜',
    '雙腳白蘿蔔': '白蘿蔔',
    '黃金白菜': '白菜',
    '結實青蔥': '青蔥',
    '扭曲牛蒡': '牛蒡',
    '黃金草莓': '草莓',
    '波浪小黃瓜': '小黃瓜',
    '黃金大蒜': '大蒜',
    '彩色玉蜀黍': '玉蜀黍',
    '巨大番茄': '番茄',
    '珠寶鳳梨': '鳳梨',
    '巨大番薯': '番薯',
    '圓球茄子': '茄子',
    '跳舞甜椒': '甜椒',
    '跳舞青椒': '青椒',
    '萬歲綠花椰菜': '綠花椰菜',
    '黃金橄欖': '橄欖',
    '黃金咖啡豆': '咖啡豆',
    '巨大酪梨': '酪梨',
    '腫包柳橙': '柳橙',
    '心形櫻桃': '櫻桃',
    '黃金香蕉': '香蕉',
    '黃金桃子': '桃子',
    '珠寶杏仁': '杏仁',
    '心形檸檬': '檸檬',
    '雞型芒果': '芒果',
    '黃金可可豆': '可可豆',
    '手指葡萄': '葡萄',
    '群花藍莓': '藍莓',
    '手指麝香葡萄': '麝香葡萄',
    '黃金蘋果': '蘋果',
}


def find_base_crop(variant_name):
    """找出變異種對應的基礎作物"""
    # 先檢查特殊對應表
    if variant_name in VARIANT_MAPPING:
        return VARIANT_MAPPING[variant_name]

    # 再檢查前綴
    for prefix in VARIANT_PREFIXES:
        if variant_name.startswith(prefix):
            return variant_name[len(prefix):]

    return None


def parse_seasons(season_str):
    """將季節字串轉換為陣列"""
    if not season_str:
        return []
    seasons = []
    season_map = {'春': '春', '夏': '夏', '秋': '秋', '冬': '冬'}
    for s in season_map:
        if s in season_str:
            seasons.append(season_map[s])
    return seasons


def parse_materials(material_str):
    """解析材料字串，分割為陣列"""
    if not material_str:
        return []
    # 以逗號或換行分割
    materials = re.split(r'[,\n]+', str(material_str))
    # 移除編號前綴（例如：1. 2. 3.）
    cleaned = []
    for m in materials:
        m = m.strip()
        if m:
            # 移除開頭的編號（數字 + 點）
            m = re.sub(r'^\d+\.\s*', '', m)
            if m:
                cleaned.append(m)
    return cleaned


def clean_value(value):
    """清理儲存格值"""
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return value
    return str(value).strip()


def parse_crops(ws):
    """解析農作物資料"""
    crops = []
    base_crops_dict = {}  # 用於變異種繼承

    # 假設第一行是標題
    headers = [cell.value for cell in ws[1]]

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:  # 跳過空行
            continue

        # Excel 欄位: 排序, 中文名稱, 日文名稱, 季節, 成長天數, 再次收成, 單次收穫量, 基本價格, 取得方式, 連作作物
        crop = {
            'name': clean_value(row[1]),  # 中文名稱
            'season': clean_value(row[3]) if len(row) > 3 else '',  # 季節
            'days': clean_value(row[4]) if len(row) > 4 else 0,  # 成長天數
            'regrow': clean_value(row[5]) if len(row) > 5 else '',  # 再次收成
            'harvest': clean_value(row[6]) if len(row) > 6 else 0,  # 單次收穫量
            'price': clean_value(row[7]) if len(row) > 7 else 0,  # 基本價格
            'continuous': clean_value(row[9]) if len(row) > 9 else '',  # 連作作物
        }

        # 解析季節陣列
        crop['seasons'] = parse_seasons(crop['season'])

        # 檢查是否為變異種
        base_name = find_base_crop(crop['name'])
        if base_name:
            # 變異種：繼承基礎資料（除了價格）
            if base_name in base_crops_dict:
                base = base_crops_dict[base_name]
                # 保留變異種自己的價格，其他繼承
                if not crop['season']:
                    crop['season'] = base['season']
                    crop['seasons'] = base['seasons']
                if not crop['days']:
                    crop['days'] = base['days']
                if not crop['regrow']:
                    crop['regrow'] = base['regrow']
                if not crop['harvest']:
                    crop['harvest'] = base['harvest']
                if not crop['continuous']:
                    crop['continuous'] = base['continuous']
        else:
            # 基礎作物，加入字典供變異種使用
            base_crops_dict[crop['name']] = crop

        crops.append(crop)

    return crops


def parse_processed(ws):
    """解析風車加工品資料"""
    processed = []

    # Excel 欄位: 排序, 風車, 中文加工品名稱, 日文名稱, 加工品類別, 材料, 基準價格, 特殊條件
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue

        # 解析材料並移除季節後綴
        materials = parse_materials(row[5]) if len(row) > 5 else []
        # 移除季節後綴 (春夏秋冬) -> 番茄
        materials = [re.sub(r'\([^)]*\)', '', m).strip() for m in materials]

        item = {
            'name': clean_value(row[2]),  # 中文加工品名稱
            'windmill': clean_value(row[1]) if len(row) > 1 else '',  # 風車
            'category': clean_value(row[4]) if len(row) > 4 else '',  # 加工品類別
            'price': clean_value(row[6]) if len(row) > 6 else 0,  # 基準價格
            'materials': materials,
        }

        processed.append(item)

    return processed


def parse_cooking(ws):
    """解析料理資料"""
    cooking = []

    # Excel 欄位: 排序, 中文名稱, 日文名稱, 類別, 基本價格, 恢復量, 效果, 廚具, 材料, ...
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue

        # 解析材料欄位（可能包含 === 巧思 === 分隔）
        material_text = clean_value(row[8]) if len(row) > 8 else ''  # 材料欄位
        materials = []
        ingenuity = []

        if '=== 巧思 ===' in material_text:
            parts = material_text.split('=== 巧思 ===')
            materials = parse_materials(parts[0])
            ingenuity = parse_materials(parts[1] if len(parts) > 1 else '')
        else:
            materials = parse_materials(material_text)

        # 移除季節後綴 (夏秋) -> 番茄
        materials = [re.sub(r'\([^)]*\)', '', m).strip() for m in materials]
        ingenuity = [re.sub(r'\([^)]*\)', '', m).strip() for m in ingenuity]

        item = {
            'name': clean_value(row[1]),  # 中文名稱
            'category': clean_value(row[3]) if len(row) > 3 else '',  # 類別
            'price': clean_value(row[4]) if len(row) > 4 else 0,  # 基本價格
            'recovery': clean_value(row[5]) if len(row) > 5 else '',  # 恢復量
            'effect': clean_value(row[6]) if len(row) > 6 else '',  # 效果
            'materials': materials,
            'ingenuity': ingenuity,
        }

        cooking.append(item)

    return cooking


def main():
    """主程式"""
    print("開始處理 Excel 資料...")

    # 讀取 Excel
    excel_path = Path('data/道具資料表.xlsx')
    if not excel_path.exists():
        print(f"❌ 找不到檔案: {excel_path}")
        return

    wb = openpyxl.load_workbook(excel_path)
    print(f"✓ 已載入 Excel 檔案")
    print(f"  工作表: {wb.sheetnames}")

    # 解析各工作表
    game_data = {}

    # 1. 農作物
    if '農作物' in wb.sheetnames:
        print("\n處理農作物資料...")
        crops = parse_crops(wb['農作物'])
        game_data['crops'] = crops
        print(f"✓ 解析了 {len(crops)} 個農作物")

        # 統計變異種
        variants = [c for c in crops if find_base_crop(c['name'])]
        print(f"  其中 {len(variants)} 個變異種")

    # 2. 風車加工品
    if '風車加工品' in wb.sheetnames:
        print("\n處理風車加工品資料...")
        processed = parse_processed(wb['風車加工品'])
        game_data['processed'] = processed
        print(f"✓ 解析了 {len(processed)} 個加工品")

    # 3. 料理
    if '料理' in wb.sheetnames:
        print("\n處理料理資料...")
        cooking = parse_cooking(wb['料理'])
        game_data['cooking'] = cooking
        print(f"✓ 解析了 {len(cooking)} 個料理")

    # 保存 JSON
    output_path = Path('data/game_data.json')
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(game_data, f, ensure_ascii=False, indent=2)

    print(f"\n✓ 成功生成 {output_path}")
    print(f"  總共 {len(game_data.get('crops', []))} + {len(game_data.get('processed', []))} + {len(game_data.get('cooking', []))} = {sum(len(v) for v in game_data.values())} 項資料")


if __name__ == '__main__':
    main()
